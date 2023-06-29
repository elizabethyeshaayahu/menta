from src.true_points import TruePoint
from src.estimated_point import EstimatePoint
import openpyxl
import numpy as np


def get_all_points(file_name: str):
    true_points = openpyxl.load_workbook(file_name)
    true_points_properties = true_points.active

    points = []
    for row in range(1, true_points_properties.max_row):
        values = []
        for col in true_points_properties.iter_cols(2, true_points_properties.max_column):
            values.append(col[row].value)

        points.append(values)

    return points


def get_all_true_points():
    true_points = []
    for row in get_all_points('../assets/true_points.xlsx'):
        true_points.append(TruePoint(*row))
    return true_points


def get_all_estimated_points():
    estimated_points = []
    for row in get_all_points('../assets/estimated_points.xlsx'):
        estimated_points.append(EstimatePoint(*row))
    return estimated_points


def to_standard_normal_distribution_point(x: float, y: float,
                                          x_mean: float, y_mean: float,
                                          x_stv: float, y_stv: float):
    return [(x - x_mean) / x_stv, (y - y_mean) / y_stv]


#k is set to 2
def chi_distribution_density_function(x: float):
    if x <= 0:
        return 0
    else:
        return 0.5*(np.exp(-x / 2))


chi_variance = 0.5 * np.sqrt(np.pi)
chi_stv = np.sqrt(chi_variance)
